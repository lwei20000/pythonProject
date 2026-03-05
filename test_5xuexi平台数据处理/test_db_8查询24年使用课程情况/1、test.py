import pandas as pd
import sys
from openpyxl import load_workbook


def analyze_courses(file_path):
    try:
        # 读取Excel文件，跳过第一行，使用第二行作为表头
        df = pd.read_excel(file_path, header=1)

        print("文件中的列名：")
        for i, col in enumerate(df.columns, 1):
            print(f"{i}. {col}")

        # 尝试自动识别课程列
        course_column = None
        possible_names = ['课程', 'course', 'Course', '科目', 'subject', 'Subject', '课程名称', '课程名']

        for col in df.columns:
            if col.strip() in possible_names:
                course_column = col
                break

        # 如果没有找到，让用户选择
        if course_column is None:
            print("\n请从上面的列名中选择课程列（输入列号）：")
            try:
                choice = int(input().strip())
                if 1 <= choice <= len(df.columns):
                    course_column = df.columns[choice - 1]
                else:
                    print("输入无效，使用第一列作为课程列")
                    course_column = df.columns[0]
            except:
                print("输入无效，使用第一列作为课程列")
                course_column = df.columns[0]

        print(f"\n使用列 '{course_column}' 作为课程列")

        # 统计每个课程的行数
        course_counts = df[course_column].value_counts()

        # 打印课程名称和对应的行数
        print("\n课程统计结果（课程名称 - 行数）：")
        print("=" * 80)
        for course, count in course_counts.items():
            print(f"{course} - {count}行")

        print("=" * 80)
        print(f"总共有 {len(course_counts)} 个不同的课程")
        print(f"总行数: {len(df)}")

        # 将统计结果保存到原始Excel的第二个sheet
        save_to_excel(file_path, course_counts)

        return course_counts

    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'")
    except Exception as e:
        print(f"处理文件时发生错误: {str(e)}")


def save_to_excel(file_path, course_counts):
    try:
        # 创建统计结果的DataFrame
        result_df = pd.DataFrame({
            '课程名称': course_counts.index,
            '行数': course_counts.values
        })

        # 按行数降序排序
        result_df = result_df.sort_values('行数', ascending=False)

        # 使用openpyxl来操作现有的Excel文件
        book = load_workbook(file_path)

        # 如果已经存在名为"课程统计"的sheet，先删除它
        if '课程统计' in book.sheetnames:
            std = book['课程统计']
            book.remove(std)

        # 创建新的sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            result_df.to_excel(writer, sheet_name='课程统计', index=False)

        print(f"\n统计结果已保存到文件: {file_path} 的第二个sheet中")

    except Exception as e:
        print(f"保存结果时发生错误: {str(e)}")
        # 如果上面的方法失败，尝试使用另一种方法
        try:
            result_df = pd.DataFrame({
                '课程名称': course_counts.index,
                '行数': course_counts.values
            })
            result_df = result_df.sort_values('行数', ascending=False)

            # 读取原始文件的所有sheet
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                result_df.to_excel(writer, sheet_name='课程统计', index=False)

            print(f"统计结果已保存到文件: {file_path} 的第二个sheet中")
        except Exception as e2:
            print(f"备用保存方法也失败: {e2}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        excel_file_path = sys.argv[1]
    else:
        excel_file_path = input("请输入Excel文件路径: ").strip()

    analyze_courses(excel_file_path)


