import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 1. 读取数据（跳过第一行说明）
df_study = pd.read_excel("用户学习进度.xlsx", header=1)
df_exam = pd.read_excel("考试信息.xlsx", header=1)

# 2. 插入新列
score_col_idx = df_study.columns.get_loc("学习成绩")
df_study.insert(score_col_idx + 1, "期末考试状态", "未参加考试")

# 3. 匹配考试状态
df_study["合并键"] = df_study["学生"] + "_" + df_study["课程"]
df_exam["合并键"] = df_exam["学生"] + "_" + df_exam["课程"]

merged = pd.merge(df_study, df_exam[["合并键", "考试状态"]],
                 on="合并键", how="left")
df_study["期末考试状态"] = merged["考试状态"].fillna("未参加考试")

# 4. 保存为临时文件（先用pandas导出）
temp_file = "临时_用户学习进度.xlsx"
df_study.drop(columns=["合并键"], inplace=True)
df_study.to_excel(temp_file, index=False)

# 5. 用openpyxl加载文件并设置单元格颜色
wb = load_workbook(temp_file)
ws = wb.active

# 定义颜色
GRAY = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 找到"期末考试状态"列
header_row = 1  # 因为header=1，所以标题在第1行
for col in range(1, ws.max_column + 1):
    if ws.cell(row=header_row, column=col).value == "期末考试状态":
        status_col = col
        break

# 遍历所有行（从第2行开始，因为第1行是标题）
for row in range(2, ws.max_row + 1):
    status = ws.cell(row=row, column=status_col).value
    if status == "未参加考试":
        ws.cell(row=row, column=status_col).fill = GRAY
    elif status == "提交未批改":
        ws.cell(row=row, column=status_col).fill = YELLOW

# 6. 保存最终文件
output_file = "用户学习进度_更新.xlsx"
wb.save(output_file)
print(f"处理完成！结果已保存为 '{output_file}'")